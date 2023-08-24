import openpyxl

#This Script reads header file, then push it to an excel sheet and give an ID for each function
 
ExcelFile = '/home/abdullah/Documents/ParsingHeader.xlsx'
HeaderFile= '/home/abdullah/Documents/Header.h'

#Function to read from excel file 
def Read_ExcelFile(path):
    data =[]
    Workbook = openpyxl.load_workbook(path)

    sheet = Workbook.active
    i = 1
    for row in sheet.iter_rows(values_only=True):
       # print(row)
        data.append(row)
            
    Workbook.close()
    return data

#Function to write to an excel sheet
def Write_ExcelFile(path,data,SavePath):
    
    Workbook = openpyxl.load_workbook(path)

    sheet = Workbook.active

    cell = 1
    #Print data for debugging
    print(data)
    
    #loop through each row in the excel file to write the function and the corresponding ID starting from ID1
    for row in data:
        print(len(row))
        
        #Write the function in coulmn B
        sheet[f'B{cell}']= " ".join(row)
        
        #write the ID in coulmn A
        sheet[f'A{cell}']= f'IDX{cell}'
        
        #Go to the next cell(row)
        cell+=1
        
    #Save the excel file and close    
    Workbook.save(SavePath)
    Workbook.close()    


#Read data from the header file
Hfile = open(HeaderFile,'r')
functions = []
for line in Hfile.readlines():

    #Check it's not an empty line
    if line.strip():
    	
    	#Conside variables and functions only MACROS are not included
        if line[0] == 'v' or line[0] == 'u' or line[0] == 's' or line[0] == 'f':
            functions.append(line.split())


Write_ExcelFile(ExcelFile,functions,ExcelFile)
